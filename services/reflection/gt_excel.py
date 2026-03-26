from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook


def extract_bundle_album_id(bundle_dir: str | Path) -> str:
    report_path = Path(bundle_dir) / "downstream_audit_report.json"
    if not report_path.exists():
        raise FileNotFoundError(f"Missing downstream audit report: {report_path}")

    payload = json.loads(report_path.read_text(encoding="utf-8"))
    album_id = _find_first_album_id(payload)
    if not album_id:
        raise ValueError(f"Could not find album_id inside {report_path}")
    return album_id


def convert_profile_gt_excel(
    *,
    workbook_path: str | Path,
    bundle_dir: str | Path,
    labeler: str,
) -> List[Dict[str, Any]]:
    album_id = extract_bundle_album_id(bundle_dir)
    actual_outputs = load_structured_profile_outputs(bundle_dir)
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    records: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        chinese_label = _stringify(row[0])
        field_key = _stringify(row[1])
        raw_result = _stringify(row[2])
        raw_confidence = _stringify(row[3])
        accuracy_note = _stringify(row[4])
        correction = _stringify(row[5])
        evidence_summary = _stringify(row[6])

        if not field_key:
            continue

        actual_output = _normalize_actual_output(actual_outputs.get(field_key), fallback=raw_result)
        gt_value = correction or actual_output
        if not gt_value:
            continue
        if not _is_usable_gt_row(accuracy_note=accuracy_note, correction=correction):
            continue

        notes_parts = [
            f"中文标签: {chinese_label}" if chinese_label else "",
            f"标注说明: {accuracy_note}" if accuracy_note else "",
            f"表格结果: {raw_result}" if raw_result and raw_result != actual_output else "",
            f"证据摘要: {evidence_summary}" if evidence_summary else "",
        ]
        if correction:
            notes_parts.append(f"人工修正: {correction}")
        record = {
            "album_id": album_id,
            "field_key": field_key,
            "gt_value": gt_value,
            "labeler": labeler,
            "notes": " | ".join(part for part in notes_parts if part),
            "source": "excel_annotation",
            "source_workbook": str(Path(workbook_path)),
            "original_output": actual_output,
            "original_confidence": raw_confidence,
            "evidence_summary": evidence_summary,
            "chinese_label": chinese_label,
            "accuracy_note": accuracy_note,
        }
        records.append(record)
    return records


def write_profile_gt_jsonl(records: Iterable[Dict[str, Any]], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def load_structured_profile_outputs(bundle_dir: str | Path) -> Dict[str, Any]:
    profile_path = Path(bundle_dir) / "structured_profile.json"
    if not profile_path.exists():
        return {}
    payload = json.loads(profile_path.read_text(encoding="utf-8"))
    flattened: Dict[str, Any] = {}
    _walk_structured_profile("", payload, flattened)
    return flattened


def _find_first_album_id(payload: Any) -> str:
    if isinstance(payload, dict):
        album_id = _stringify(payload.get("album_id"))
        if album_id:
            return album_id
        for value in payload.values():
            nested = _find_first_album_id(value)
            if nested:
                return nested
        return ""
    if isinstance(payload, list):
        for item in payload:
            nested = _find_first_album_id(item)
            if nested:
                return nested
    return ""


def _walk_structured_profile(prefix: str, payload: Any, flattened: Dict[str, Any]) -> None:
    if isinstance(payload, dict) and {"value", "confidence", "evidence", "reasoning"}.issubset(payload.keys()):
        flattened[prefix] = payload.get("value")
        return
    if isinstance(payload, dict):
        for key, value in payload.items():
            child_prefix = f"{prefix}.{key}" if prefix else key
            _walk_structured_profile(child_prefix, value, flattened)


def _is_usable_gt_row(*, accuracy_note: str, correction: str) -> bool:
    if correction:
        return True
    return accuracy_note.startswith("1")


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_actual_output(value: Any, *, fallback: str) -> str:
    if value in (None, "", []):
        return fallback
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip()
